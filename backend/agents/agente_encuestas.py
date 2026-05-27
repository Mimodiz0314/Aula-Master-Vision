"""
Subagente de Encuestas
=======================
Gestión avanzada del ciclo de vida: filtros, métricas, copia y edición.
"""
from sqlalchemy.orm import Session
from sqlalchemy import func
from models.db_models import Evaluacion, Respuesta
from typing import Optional
import json


def obtener_stats_docente(db: Session, docente_id: int) -> dict:
    """Calcula estadísticas reales para el panel del docente."""
    total = db.query(Evaluacion).filter(Evaluacion.docente_id == docente_id).count()
    activas = db.query(Evaluacion).filter(
        Evaluacion.docente_id == docente_id,
        Evaluacion.activa == True
    ).count()
    # Total respuestas de todas las evaluaciones del docente
    total_respuestas = (
        db.query(func.count(Respuesta.id))
        .join(Evaluacion, Respuesta.evaluacion_id == Evaluacion.id)
        .filter(Evaluacion.docente_id == docente_id)
        .scalar() or 0
    )
    return {
        "total_evaluaciones": total,
        "activas": activas,
        "completadas": total - activas,
        "total_respuestas": total_respuestas,
    }


def editar_evaluacion(db: Session, evaluacion_id: int, titulo: str, descripcion: str,
                      es_anonima: bool, preguntas: list) -> Optional[Evaluacion]:
    """Edita una evaluación existente."""
    ev = db.query(Evaluacion).filter(Evaluacion.id == evaluacion_id).first()
    if not ev:
        return None
    ev.titulo = titulo
    ev.descripcion = descripcion
    ev.es_anonima = es_anonima
    ev.preguntas = preguntas
    db.commit()
    db.refresh(ev)
    return ev


def clonar_evaluacion(db: Session, evaluacion_id: int, docente_id: int) -> Optional[Evaluacion]:
    """Clona una evaluación con un nuevo código de acceso."""
    import random, string
    ev_orig = db.query(Evaluacion).filter(Evaluacion.id == evaluacion_id).first()
    if not ev_orig:
        return None

    def _gen_codigo():
        return "AULA-" + "".join(random.choices(string.ascii_uppercase + string.digits, k=8))

    nuevo_codigo = _gen_codigo()
    while db.query(Evaluacion).filter(Evaluacion.codigo_acceso == nuevo_codigo).first():
        nuevo_codigo = _gen_codigo()

    copia = Evaluacion(
        titulo=f"[Copia] {ev_orig.titulo}",
        descripcion=ev_orig.descripcion,
        es_anonima=ev_orig.es_anonima,
        preguntas=ev_orig.preguntas,
        docente_id=docente_id,
        codigo_acceso=nuevo_codigo,
        activa=False,
    )
    db.add(copia)
    db.commit()
    db.refresh(copia)
    return copia


def eliminar_evaluacion(db: Session, evaluacion_id: int) -> bool:
    """Elimina una evaluación y sus respuestas."""
    ev = db.query(Evaluacion).filter(Evaluacion.id == evaluacion_id).first()
    if not ev:
        return False
    # Eliminar respuestas asociadas
    db.query(Respuesta).filter(Respuesta.evaluacion_id == evaluacion_id).delete()
    db.delete(ev)
    db.commit()
    return True


def listar_respuestas_evaluacion(db: Session, evaluacion_id: int) -> list:
    """Retorna todas las respuestas de una evaluación con análisis IA."""
    respuestas = db.query(Respuesta).filter(Respuesta.evaluacion_id == evaluacion_id).all()
    resultado = []
    for r in respuestas:
        analisis = []
        if r.analisis_ia:
            try:
                analisis = json.loads(r.analisis_ia)
            except Exception:
                pass
        resultado.append({
            "id": r.id,
            "nombre": r.nombre_estudiante or "Anónimo",
            "grado": r.grado or "—",
            "datos": r.datos,
            "analisis_ia": analisis,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        })
    return resultado


def exportar_excel(db: Session, evaluacion_id: int) -> bytes:
    """
    Genera un archivo Excel (.xlsx) profesional con 3 hojas:
      1. Respuestas    — tabla completa de todas las respuestas
      2. Estadísticas  — promedios, distribución y sentimientos
      3. Análisis IA   — resumen semántico por participante (Groq)
    """
    import io, json
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    ev = db.query(Evaluacion).filter(Evaluacion.id == evaluacion_id).first()
    if not ev:
        return b""

    respuestas = db.query(Respuesta).filter(Respuesta.evaluacion_id == evaluacion_id).all()
    preguntas  = ev.preguntas or []

    # ── Paleta de colores ──────────────────────────────────────
    TEAL_DARK   = "0F766E"   # encabezados principales
    TEAL_MED    = "14B8A6"   # encabezados secundarios
    TEAL_LIGHT  = "CCFBF1"   # filas alternas
    GRAY_LIGHT  = "F8FAFC"   # filas alternas 2
    WHITE       = "FFFFFF"
    GOLD        = "F59E0B"
    GREEN       = "22C55E"
    RED_SOFT    = "FEE2E2"
    GREEN_SOFT  = "DCFCE7"
    AMBER_SOFT  = "FEF3C7"

    def fill(hex_color: str):
        return PatternFill("solid", fgColor=hex_color)

    def bold_font(size=11, color="000000", bold=True):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def border_thin():
        s = Side(style="thin", color="E2E8F0")
        return Border(left=s, right=s, top=s, bottom=s)

    def set_col_width(ws, col_idx: int, width: float):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # HOJA 1 — Respuestas
    # ════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Respuestas"
    ws1.sheet_view.showGridLines = False

    # Fila 1: título de la evaluación
    total_cols = 4 + len(preguntas)
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws1.cell(1, 1, f"📊  {ev.titulo}")
    c.fill      = fill(TEAL_DARK)
    c.font      = bold_font(14, WHITE)
    c.alignment = center()
    ws1.row_dimensions[1].height = 32

    # Fila 2: metadatos
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    fecha_gen = datetime.now().strftime("%d/%m/%Y %H:%M")
    tipo_ev   = "🔒 Anónima" if ev.es_anonima else "👤 Nominal"
    meta = f"Generado: {fecha_gen}   |   Tipo: {tipo_ev}   |   Total respuestas: {len(respuestas)}"
    c2 = ws1.cell(2, 1, meta)
    c2.fill      = fill(TEAL_MED)
    c2.font      = bold_font(10, WHITE, bold=False)
    c2.alignment = center()
    ws1.row_dimensions[2].height = 20

    # Fila 3: encabezados de columnas
    headers = ["#", "Nombre", "Grado", "Fecha"] + [
        f"P{p['id']} — {p['texto'][:50]}" for p in preguntas
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(3, col_idx, h)
        cell.fill      = fill("1E293B")
        cell.font      = bold_font(10, WHITE)
        cell.alignment = center()
        cell.border    = border_thin()
    ws1.row_dimensions[3].height = 36

    # Filas de datos
    for row_idx, r in enumerate(respuestas, 4):
        analisis_ia = []
        if r.analisis_ia:
            try: analisis_ia = json.loads(r.analisis_ia)
            except: pass

        bg = TEAL_LIGHT if row_idx % 2 == 0 else GRAY_LIGHT
        datos_row = [
            row_idx - 3,
            r.nombre_estudiante or "Anónimo",
            r.grado or "—",
            r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else "—",
        ]
        for p in preguntas:
            val = (r.datos or {}).get(str(p["id"]), "")
            if p.get("tipo") == "escala" and val:
                try:
                    n = int(val)
                    bloques = "█" * n + "░" * (10 - n)
                    val = f"{n}/10  {bloques}"
                except:
                    pass
            datos_row.append(val)

        for col_idx, val in enumerate(datos_row, 1):
            cell = ws1.cell(row_idx, col_idx, val)
            cell.fill      = fill(bg)
            cell.border    = border_thin()
            cell.alignment = center() if col_idx <= 4 else left()
            cell.font      = Font(name="Calibri", size=10)

        ws1.row_dimensions[row_idx].height = 18

    # Anchos de columna
    set_col_width(ws1, 1, 5)
    set_col_width(ws1, 2, 22)
    set_col_width(ws1, 3, 14)
    set_col_width(ws1, 4, 18)
    for i in range(5, total_cols + 1):
        p = preguntas[i - 5]
        set_col_width(ws1, i, 45 if p.get("tipo") == "abierta" else 22)

    # ════════════════════════════════════════════════════════════
    # HOJA 2 — Estadísticas
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Estadísticas")
    ws2.sheet_view.showGridLines = False

    # Título
    ws2.merge_cells("A1:F1")
    t = ws2.cell(1, 1, "📈  Estadísticas de la Evaluación")
    t.fill = fill(TEAL_DARK); t.font = bold_font(13, WHITE); t.alignment = center()
    ws2.row_dimensions[1].height = 30

    # Resumen general
    ws2.merge_cells("A3:F3")
    ws2.cell(3, 1, "RESUMEN GENERAL").fill = fill("1E293B")
    ws2.cell(3, 1).font = bold_font(11, WHITE); ws2.cell(3, 1).alignment = center()

    resumen = [
        ("Evaluación",       ev.titulo),
        ("Total respuestas", len(respuestas)),
        ("Tipo",             "Anónima" if ev.es_anonima else "Nominal"),
        ("Fecha creación",   ev.created_at.strftime("%d/%m/%Y") if ev.created_at else "—"),
    ]
    for i, (label, val) in enumerate(resumen, 4):
        ws2.cell(i, 1, label).font  = bold_font(10)
        ws2.cell(i, 1).fill         = fill(TEAL_LIGHT)
        ws2.cell(i, 1).alignment    = left()
        ws2.cell(i, 2, val).font    = Font(name="Calibri", size=10)
        ws2.cell(i, 2).alignment    = left()
        ws2.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    # Promedios por pregunta de escala
    preguntas_escala = [p for p in preguntas if p.get("tipo") == "escala"]
    if preguntas_escala:
        row = len(resumen) + 6
        ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws2.cell(row, 1, "PROMEDIOS POR PREGUNTA (ESCALA 1-10)").fill = fill("1E293B")
        ws2.cell(row, 1).font = bold_font(11, WHITE); ws2.cell(row, 1).alignment = center()
        row += 1

        hdrs = ["Pregunta", "Texto", "Promedio", "Mín", "Máx", "Distribución"]
        for ci, h in enumerate(hdrs, 1):
            cell = ws2.cell(row, ci, h)
            cell.fill = fill(TEAL_MED); cell.font = bold_font(10, WHITE); cell.alignment = center()
        row += 1

        for p in preguntas_escala:
            valores = []
            for r in respuestas:
                v = (r.datos or {}).get(str(p["id"]))
                if v:
                    try: valores.append(int(v))
                    except: pass

            if valores:
                promedio = round(sum(valores) / len(valores), 2)
                distrib  = " | ".join([f"{n}:{'▮'*valores.count(n)}" for n in range(1,11) if valores.count(n)])

                prom_color = GREEN if promedio >= 7 else GOLD if promedio >= 5 else "EF4444"
                data_row = [f"P{p['id']}", p["texto"][:60], promedio, min(valores), max(valores), distrib]
                for ci, val in enumerate(data_row, 1):
                    cell = ws2.cell(row, ci, val)
                    cell.border    = border_thin()
                    cell.alignment = center() if ci != 2 else left()
                    cell.font      = Font(name="Calibri", size=10)
                    if ci == 3:
                        cell.fill = fill(GREEN_SOFT if promedio >= 7 else AMBER_SOFT if promedio >= 5 else RED_SOFT)
                        cell.font = bold_font(10, "166534" if promedio >= 7 else "92400E" if promedio >= 5 else "991B1B")
                row += 1

    # Distribución de sentimientos
    row += 2
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws2.cell(row, 1, "DISTRIBUCIÓN DE SENTIMIENTOS (IA GROQ)").fill = fill("1E293B")
    ws2.cell(row, 1).font = bold_font(11, WHITE); ws2.cell(row, 1).alignment = center()
    row += 1

    sentimientos = {"positivo": 0, "neutro": 0, "negativo": 0}
    for r in respuestas:
        if r.analisis_ia:
            try:
                ia = json.loads(r.analisis_ia)
                s  = ia[0].get("sentimiento", "neutro") if ia else "neutro"
                sentimientos[s] = sentimientos.get(s, 0) + 1
            except: pass

    sent_config = [
        ("😊 Positivo", sentimientos["positivo"], GREEN_SOFT,  "166534"),
        ("😐 Neutro",   sentimientos["neutro"],   AMBER_SOFT,  "92400E"),
        ("😟 Negativo", sentimientos["negativo"], RED_SOFT,    "991B1B"),
    ]
    for label, cnt, bg, fg in sent_config:
        pct = f"{round(cnt/len(respuestas)*100)}%" if respuestas else "0%"
        for ci, val in enumerate([label, cnt, pct], 1):
            cell = ws2.cell(row, ci, val)
            cell.fill = fill(bg); cell.font = bold_font(10, fg)
            cell.alignment = center(); cell.border = border_thin()
        row += 1

    # Anchos hoja 2
    for ci, w in [(1,12),(2,55),(3,12),(4,8),(5,8),(6,35)]:
        set_col_width(ws2, ci, w)

    # ════════════════════════════════════════════════════════════
    # HOJA 3 — Análisis IA
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Análisis IA")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    t3 = ws3.cell(1, 1, "🧠  Análisis Semántico — Groq IA")
    t3.fill = fill(TEAL_DARK); t3.font = bold_font(13, WHITE); t3.alignment = center()
    ws3.row_dimensions[1].height = 30

    ws3.merge_cells("A2:F2")
    ws3.cell(2, 1, "Análisis individual de respuestas abiertas procesadas por Inteligencia Artificial")
    ws3.cell(2, 1).fill = fill(TEAL_MED); ws3.cell(2, 1).font = bold_font(10, WHITE, False)
    ws3.cell(2, 1).alignment = center()

    hdrs3 = ["Nombre", "Grado", "Sentimiento", "Comprensión /10", "Palabras Clave", "Resumen IA"]
    for ci, h in enumerate(hdrs3, 1):
        cell = ws3.cell(3, ci, h)
        cell.fill = fill("1E293B"); cell.font = bold_font(10, WHITE); cell.alignment = center()
    ws3.row_dimensions[3].height = 28

    sent_fills = {
        "positivo": (GREEN_SOFT,  "166534"),
        "neutro":   (AMBER_SOFT,  "92400E"),
        "negativo": (RED_SOFT,    "991B1B"),
    }
    for row_idx, r in enumerate(respuestas, 4):
        analisis_ia = []
        if r.analisis_ia:
            try: analisis_ia = json.loads(r.analisis_ia)
            except: pass

        ia = analisis_ia[0] if analisis_ia else {}
        sent      = ia.get("sentimiento", "—")
        puntaje   = ia.get("puntaje_comprension", "—")
        palabras  = ", ".join(ia.get("palabras_clave", [])) if ia.get("palabras_clave") else "—"
        resumen   = ia.get("resumen", "Sin análisis disponible.")

        bg = TEAL_LIGHT if row_idx % 2 == 0 else GRAY_LIGHT
        datos3 = [
            r.nombre_estudiante or "Anónimo",
            r.grado or "—",
            sent.capitalize() if sent != "—" else "—",
            puntaje,
            palabras,
            resumen,
        ]
        for ci, val in enumerate(datos3, 1):
            cell = ws3.cell(row_idx, ci, val)
            cell.border    = border_thin()
            cell.alignment = center() if ci < 4 else left()
            cell.font      = Font(name="Calibri", size=10)
            if ci == 3 and sent in sent_fills:
                sbg, sfg = sent_fills[sent]
                cell.fill = fill(sbg); cell.font = bold_font(10, sfg)
            elif ci == 4 and isinstance(puntaje, (int, float)):
                p_color = GREEN_SOFT if puntaje >= 7 else AMBER_SOFT if puntaje >= 5 else RED_SOFT
                cell.fill = fill(p_color)
            else:
                cell.fill = fill(bg)
        ws3.row_dimensions[row_idx].height = 20

    for ci, w in [(1,22),(2,14),(3,14),(4,16),(5,30),(6,60)]:
        set_col_width(ws3, ci, w)

    # ── Guardar en buffer y retornar bytes ────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

    return output.getvalue()
